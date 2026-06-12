from flask import Flask, request, jsonify, render_template
import os
import psycopg2
import psycopg2.extras
import unicodedata
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
STOCK_MINIMO = 10
STOCK_CRITICO = 5
DIAS_ALERTA_VENCIMIENTO = 30
DATE_FMT = "%Y-%m-%d"
FMT_FECHA_REPORTE = "%Y%m%d"
PREMIUM_ENABLED = False

MARGEN_POR_CATEGORIA = {
    "Analgésicos / Antiinflamatorios": 0.30,
    "Antibióticos": 0.25,
    "Antialérgicos": 0.35,
    "Antidiabéticos / Antihipertensivos": 0.22,
    "Gastrointestinal": 0.32,
    "Dermatológicos": 0.45,
    "Suplementos / Vitaminas": 0.40,
    "Otros": 0.28,
    "Higiene Bucal": 0.35,
    "Higiene Personal": 0.40,
    "Cuidado Capilar": 0.45,
    "Cuidado de la Piel": 0.50,
    "Salud Sexual": 0.40,
    "Cuidado del Bebe": 0.30,
    "Primeros Auxilios": 0.35,
    "Dispositivos Medicos": 0.25,
    "Descongestionantes": 0.30,
    "Antiparasitarios": 0.30,
    "Hormonales": 0.25,
    "Oncología": 0.20,
    "Neurológicos": 0.25,
    "Urológicos": 0.30,
    "Osteoporosis": 0.25,
    "Oftalmológicos": 0.35,
}

# ── HELPER DE QUERIES ────────────────────────────────────
# ¿Por qué este helper existe?
# ------------------------------------------------------------
# sqlite3 permitía hacer `conn.execute("SELECT ...").fetchone()` directamente.
# psycopg2 NO: hay que hacer `cur = conn.cursor(); cur.execute(...)`.
#
# En vez de ensuciar cada función con `cur = conn.cursor(); cur.execute(...)`
# (opción repetitiva y verbosa), centralizamos el patrón en este helper.
#
# Si en el futuro cambias psycopg2 por SQLAlchemy u otra librería, solo
# cambias ESTA función, no las 38 llamadas que la usan.
#
# Uso:
#   total = query(conn, "SELECT SUM(cantidad) as t FROM inventario").fetchone()["t"]
#   rows  = query(conn, "SELECT * FROM catalogo").fetchall()
#   query(conn, "DELETE FROM inventario WHERE id = %s", (item_id,))  # INSERT/UPDATE/DELETE
def query(conn, sql, params=None):
    cur = conn.cursor()
    cur.execute(sql, params)
    return cur

def get_db():
    url = os.environ.get("DATABASE_URL")
    if not url:
        raise RuntimeError(
            "DATABASE_URL no está configurada. "
            "Defínela con: set DATABASE_URL=postgresql://usuario:clave@localhost:5432/farmacia"
        )
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    return psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)

def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS catalogo (
            id          SERIAL PRIMARY KEY,
            nombre      TEXT NOT NULL,
            principio   TEXT,
            laboratorio TEXT,
            descripcion TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS inventario (
            id               SERIAL PRIMARY KEY,
            invima_id        INTEGER,
            catalogo_id      INTEGER,
            nombre           TEXT NOT NULL,
            principio        TEXT,
            laboratorio      TEXT,
            registro         TEXT,
            cantidad         INTEGER NOT NULL DEFAULT 0,
            precio           DOUBLE PRECISION NOT NULL DEFAULT 0,
            fecha_vencimiento TEXT,
            lote             TEXT
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS ventas (
            id              SERIAL PRIMARY KEY,
            inventario_id   INTEGER NOT NULL,
            nombre          TEXT NOT NULL,
            laboratorio     TEXT,
            cantidad        INTEGER NOT NULL,
            precio_unitario DOUBLE PRECISION NOT NULL,
            total           DOUBLE PRECISION NOT NULL,
            fecha           TEXT NOT NULL
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invima (
            id               SERIAL PRIMARY KEY,
            expediente       TEXT,
            producto         TEXT,
            principio_activo TEXT,
            registro         TEXT,
            estado           TEXT,
            modalidad        TEXT,
            titular          TEXT
        )
    """)

    _ensure_column(conn, "inventario", "fecha_vencimiento", "TEXT")
    _ensure_column(conn, "inventario", "lote", "TEXT")
    _ensure_column(conn, "inventario", "categoria", "TEXT")

    conn.commit()
    conn.close()


def _ensure_column(conn, table, column, definition):
    cur = conn.cursor()
    cur.execute("""
        SELECT 1 FROM information_schema.columns
        WHERE table_schema = 'public'
          AND table_name = %s
          AND column_name = %s
    """, (table, column))
    if not cur.fetchone():
        cur.execute(f'ALTER TABLE "{table}" ADD COLUMN "{column}" {definition}')

# ── INVIMA ────
@app.route("/api/invima/buscar")
def invima_buscar():
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])
    conn = get_db()
    rows = query(conn, """
        SELECT id, expediente, producto, principio_activo, registro, modalidad, titular
        FROM invima
        WHERE producto ILIKE %s OR principio_activo ILIKE %s OR titular ILIKE %s
        LIMIT 30
    """, (f"%{q}%", f"%{q}%", f"%{q}%")).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

CATEGORIAS_PREDEFINIDAS = [
    "Analgésicos / Antiinflamatorios", "Antibióticos", "Antialérgicos",
    "Antidiabéticos / Antihipertensivos", "Gastrointestinal", "Dermatológicos",
    "Suplementos / Vitaminas", "Otros",
    "Higiene Bucal", "Higiene Personal", "Cuidado Capilar",
    "Cuidado de la Piel", "Salud Sexual", "Cuidado del Bebe",
    "Primeros Auxilios", "Dispositivos Medicos", "Descongestionantes",
    "Antiparasitarios", "Hormonales", "Oncología", "Neurológicos",
    "Urológicos", "Osteoporosis", "Oftalmológicos",
]

CATEGORIAS_KEYWORDS = {
    "ibuprofeno": "Analgésicos / Antiinflamatorios", "diclofenaco": "Analgésicos / Antiinflamatorios",
    "naproxeno": "Analgésicos / Antiinflamatorios", "ketorolaco": "Analgésicos / Antiinflamatorios",
    "paracetamol": "Analgésicos / Antiinflamatorios", "acetaminofen": "Analgésicos / Antiinflamatorios",
    "tramadol": "Analgésicos / Antiinflamatorios", "morfina": "Analgésicos / Antiinflamatorios",
    "codeina": "Analgésicos / Antiinflamatorios", "aspirina": "Analgésicos / Antiinflamatorios",
    "metamizol": "Analgésicos / Antiinflamatorios", "piroxicam": "Analgésicos / Antiinflamatorios",
    "celecoxib": "Analgésicos / Antiinflamatorios", "meloxicam": "Analgésicos / Antiinflamatorios",
    "flurbiprofeno": "Analgésicos / Antiinflamatorios", "nimesulida": "Analgésicos / Antiinflamatorios",
    "indometacina": "Analgésicos / Antiinflamatorios", "fenilbutazona": "Analgésicos / Antiinflamatorios",
    "amoxicilina": "Antibióticos", "azitromicina": "Antibióticos", "ciprofloxacino": "Antibióticos",
    "ceftriaxona": "Antibióticos", "amoxicilina clavulanato": "Antibióticos",
    "metronidazol": "Antibióticos", "doxiciclina": "Antibióticos", "penicilina": "Antibióticos",
    "cefuroxima": "Antibióticos", "levofloxacino": "Antibióticos", "cefalexina": "Antibióticos",
    "sulfametoxazol": "Antibióticos", "trimetoprim": "Antibióticos", "clindamicina": "Antibióticos",
    "eritromicina": "Antibióticos", "vancomicina": "Antibióticos", "gentamicina": "Antibióticos",
    "tobramicina": "Antibióticos", "cefepima": "Antibióticos", "meropenem": "Antibióticos",
    "piperacilina": "Antibióticos", "tazobactam": "Antibióticos", "linezolid": "Antibióticos",
    "rifampicina": "Antibióticos", "isoniazida": "Antibióticos", "pirazinamida": "Antibióticos",
    "etambutol": "Antibióticos", "fluconazol": "Antibióticos", "voriconazol": "Antibióticos",
    "itraconazol": "Antibióticos", "anfotericina": "Antibióticos", "acyclovir": "Antibióticos",
    "valacyclovir": "Antibióticos", "oseltamivir": "Antibióticos",
    "loratadina": "Antialérgicos", "cetirizina": "Antialérgicos", "desloratadina": "Antialérgicos",
    "hidroxizina": "Antialérgicos", "clorfeniramina": "Antialérgicos", "difenhidramina": "Antialérgicos",
    "fexofenadina": "Antialérgicos", "levocetirizina": "Antialérgicos",
    "metformina": "Antidiabéticos / Antihipertensivos", "glibenclamida": "Antidiabéticos / Antihipertensivos",
    "insulina": "Antidiabéticos / Antihipertensivos", "lisinopril": "Antidiabéticos / Antihipertensivos",
    "enalapril": "Antidiabéticos / Antihipertensivos", "losartan": "Antidiabéticos / Antihipertensivos",
    "amlodipino": "Antidiabéticos / Antihipertensivos", "hidroclorotiazida": "Antidiabéticos / Antihipertensivos",
    "atorvastatin": "Antidiabéticos / Antihipertensivos", "simvastatin": "Antidiabéticos / Antihipertensivos",
    "valsartan": "Antidiabéticos / Antihipertensivos", "irbesartan": "Antidiabéticos / Antihipertensivos",
    "ramipril": "Antidiabéticos / Antihipertensivos", "perindopril": "Antidiabéticos / Antihipertensivos",
    "bisoprolol": "Antidiabéticos / Antihipertensivos", "metoprolol": "Antidiabéticos / Antihipertensivos",
    "propranolol": "Antidiabéticos / Antihipertensivos", "atenolol": "Antidiabéticos / Antihipertensivos",
    "carvedilol": "Antidiabéticos / Antihipertensivos", "espironolactona": "Antidiabéticos / Antihipertensivos",
    "furosemida": "Antidiabéticos / Antihipertensivos", "torsemida": "Antidiabéticos / Antihipertensivos",
    "nitroglicerina": "Antidiabéticos / Antihipertensivos",
    "nifedipino": "Antidiabéticos / Antihipertensivos", "verapamilo": "Antidiabéticos / Antihipertensivos",
    "diltiazem": "Antidiabéticos / Antihipertensivos", "clopidogrel": "Antidiabéticos / Antihipertensivos",
    "warfarina": "Antidiabéticos / Antihipertensivos", "heparina": "Antidiabéticos / Antihipertensivos",
    "enoxaparina": "Antidiabéticos / Antihipertensivos", "rivaroxaban": "Antidiabéticos / Antihipertensivos",
    "apixaban": "Antidiabéticos / Antihipertensivos", "dabigatran": "Antidiabéticos / Antihipertensivos",
    "omeprazol": "Gastrointestinal", "pantoprazol": "Gastrointestinal", "esomeprazol": "Gastrointestinal",
    "ranitidina": "Gastrointestinal", "famotidina": "Gastrointestinal", "metoclopramida": "Gastrointestinal",
    "domperidona": "Gastrointestinal", "loperamida": "Gastrointestinal", "buscapina": "Gastrointestinal",
    "lansoprazol": "Gastrointestinal", "sucralfato": "Gastrointestinal", "misoprostol": "Gastrointestinal",
    "bismuto": "Gastrointestinal", "simeticona": "Gastrointestinal", "lactulosa": "Gastrointestinal",
    "bisacodilo": "Gastrointestinal", "polietilenglicol": "Gastrointestinal",
    "hidrocortisona": "Dermatológicos", "betametasona": "Dermatológicos", "clotrimazol": "Dermatológicos",
    "miconazol": "Dermatológicos", "aceite mineral": "Dermatológicos", "urea": "Dermatológicos",
    "acido retinoico": "Dermatológicos", "retinol": "Dermatológicos", "tretinoina": "Dermatológicos",
    "adapaleno": "Dermatológicos", "peróxido de benzoilo": "Dermatológicos",
    "neomicina": "Dermatológicos", "bacitracina": "Dermatológicos", "mupirocina": "Dermatológicos",
    "pimecrolimus": "Dermatológicos", "tacrolimus": "Dermatológicos",
    "metotrexato": "Dermatológicos",
    "vitamina": "Suplementos / Vitaminas", "magnesio": "Suplementos / Vitaminas", "zinc": "Suplementos / Vitaminas",
    "hierro": "Suplementos / Vitaminas", "calcio": "Suplementos / Vitaminas", "vitamina d": "Suplementos / Vitaminas",
    "vitamina c": "Suplementos / Vitaminas",
    "acido folico": "Suplementos / Vitaminas", "folato": "Suplementos / Vitaminas",
    "vitamina b12": "Suplementos / Vitaminas", "vitamina b6": "Suplementos / Vitaminas",
    "vitamina a": "Suplementos / Vitaminas", "vitamina e": "Suplementos / Vitaminas",
    "selenio": "Suplementos / Vitaminas", "cromo": "Suplementos / Vitaminas",
    "potasio": "Suplementos / Vitaminas", "flúor": "Suplementos / Vitaminas",
    "shampoo": "Cuidado Capilar", "acondicionador": "Cuidado Capilar", "gel capilar": "Cuidado Capilar",
    "mascarilla capilar": "Cuidado Capilar", "crema de manos": "Cuidado de la Piel",
    "protector solar": "Cuidado de la Piel", "crema facial": "Cuidado de la Piel",
    "bálsamo labial": "Cuidado de la Piel", "loción": "Cuidado de la Piel",
    "crema corporal": "Cuidado de la Piel", "crema hidratante": "Cuidado de la Piel",
    "pasta dental": "Higiene Bucal", "cepillo dental": "Higiene Bucal",
    "enjuague bucal": "Higiene Bucal", "hilo dental": "Higiene Bucal",
    "desodorante": "Higiene Personal", "antitranspirante": "Higiene Personal",
    "jabón": "Higiene Personal", "jabon": "Higiene Personal",
    "condón": "Salud Sexual", "condon": "Salud Sexual", "preservativo": "Salud Sexual",
    "lubricante": "Salud Sexual",
    "spray nasal": "Descongestionantes", "descongestionante": "Descongestionantes",
    "guaifenesina": "Descongestionantes", "pseudoefedrina": "Descongestionantes",
    "fenilefrina": "Descongestionantes", "oximetazolina": "Descongestionantes",
    "montelukast": "Descongestionantes", "budesonida": "Descongestionantes",
    "salbutamol": "Descongestionantes", "ipratropio": "Descongestionantes",
    "teofilina": "Descongestionantes", "formoterol": "Descongestionantes",
    "omega": "Suplementos / Vitaminas", "colágen": "Suplementos / Vitaminas", "biotina": "Suplementos / Vitaminas",
    "complejo b": "Suplementos / Vitaminas",
    "sales": "Suplementos / Vitaminas",
    "benznidazol": "Antiparasitarios",
    "albendazol": "Antiparasitarios", "mebendazol": "Antiparasitarios",
    "ivermectina": "Antiparasitarios", "prazicuantel": "Antiparasitarios",
    "nifurtimox": "Antiparasitarios",
    "levotiroxina": "Hormonales", "tiroxina": "Hormonales",
    "estrógeno": "Hormonales", "progesterona": "Hormonales",
    "testosterona": "Hormonales", "tamoxifeno": "Hormonales",
    "clomifeno": "Hormonales", "danazol": "Hormonales",
    "sitagliptina": "Antidiabéticos / Antihipertensivos", "saxagliptina": "Antidiabéticos / Antihipertensivos",
    "liraglutida": "Antidiabéticos / Antihipertensivos", "semaglutida": "Antidiabéticos / Antihipertensivos",
    "empagliflozina": "Antidiabéticos / Antihipertensivos", "dapagliflozina": "Antidiabéticos / Antihipertensivos",
    "canagliflozina": "Antidiabéticos / Antihipertensivos",
    "pioglitazona": "Antidiabéticos / Antihipertensivos", "rosiglitazona": "Antidiabéticos / Antihipertensivos",
    "glimepirida": "Antidiabéticos / Antihipertensivos", "glipizida": "Antidiabéticos / Antihipertensivos",
    "paclitaxel": "Oncología", "docetaxel": "Oncología", "doxorubicina": "Oncología",
    "ciclofosfamida": "Oncología", "fluorouracilo": "Oncología", "capecitabina": "Oncología",
    "cisplatino": "Oncología", "carboplatino": "Oncología", "oxaliplatino": "Oncología",
    "bevacizumab": "Oncología", "trastuzumab": "Oncología", "rituximab": "Oncología",
    "pembrolizumab": "Oncología", "nivolumab": "Oncología", "atezolizumab": "Oncología",
    "imatinib": "Oncología", "erlotinib": "Oncología", "gefitinib": "Oncología",
    "letrozol": "Oncología", "anastrozol": "Oncología", "exemestano": "Oncología",
    "bicalutamida": "Oncología", "leuprolide": "Oncología", "goserelina": "Oncología",
    "abiraterona": "Oncología", "enzalutamida": "Oncología",
    "raloxifeno": "Oncología",
    "pamidronato": "Oncología", "zoledronico": "Oncología",
    "azatioprina": "Oncología", "micofenolato": "Oncología",
    "ciclosporina": "Oncología",
    "dexametasona": "Oncología", "prednisona": "Oncología", "prednisolona": "Oncología",
    "metilprednisolona": "Oncología",
    "levodopa": "Neurológicos", "carbidopa": "Neurológicos",
    "bromocriptina": "Neurológicos", "pramipexol": "Neurológicos", "ropinirol": "Neurológicos",
    "donepezilo": "Neurológicos", "rivastigmina": "Neurológicos", "memantina": "Neurológicos",
    "carbamazepina": "Neurológicos", "valproico": "Neurológicos", "lamotrigina": "Neurológicos",
    "levetiracetam": "Neurológicos", "fenitoina": "Neurológicos", "topiramato": "Neurológicos",
    "gabapentina": "Neurológicos", "pregabalina": "Neurológicos",
    "duloxetina": "Neurológicos", "venlafaxina": "Neurológicos", "amitriptilina": "Neurológicos",
    "fluoxetina": "Neurológicos", "sertralina": "Neurológicos", "escitalopram": "Neurológicos",
    "citalopram": "Neurológicos", "paroxetina": "Neurológicos",
    "alprazolam": "Neurológicos", "lorazepam": "Neurológicos", "diazepam": "Neurológicos",
    "clonazepam": "Neurológicos", "zolpidem": "Neurológicos",
    "quetiapina": "Neurológicos", "olanzapina": "Neurológicos", "risperidona": "Neurológicos",
    "aripiprazol": "Neurológicos", "haloperidol": "Neurológicos", "clozapina": "Neurológicos",
    "litio": "Neurológicos",
    "tadalafil": "Urológicos", "sildenafilo": "Urológicos", "finasterida": "Urológicos",
    "dutasterida": "Urológicos", "tamsulosina": "Urológicos", "silodosina": "Urológicos",
    "solifenacina": "Urológicos", "tolterodina": "Urológicos", "oxibutinina": "Urológicos",
    "dapoxetina": "Urológicos",
    "alendronato": "Osteoporosis", "risedronato": "Osteoporosis", "etidronato": "Osteoporosis",
    "calcitonina": "Osteoporosis", "denosumab": "Osteoporosis", "teriparatida": "Osteoporosis",
    "alopurinol": "Otros", "colchicina": "Otros", "benzbromarona": "Otros",
    "ácido úrico": "Otros", "ácido urico": "Otros",
    "pilocarpina": "Oftalmológicos", "latanoprost": "Oftalmológicos", "timolol": "Oftalmológicos",
    "brimonidina": "Oftalmológicos", "dorzolamida": "Oftalmológicos",
    "artificial lagrimas": "Oftalmológicos", "lacrimal": "Oftalmológicos",
}

def _normalizar(texto):
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def inferir_categoria(principio, nombre=""):
    campos = _normalizar(f"{principio or ''} {nombre or ''}").lower().strip()
    if not campos:
        return ""
    for kw, cat in CATEGORIAS_KEYWORDS.items():
        if kw in campos:
            return cat
    return "Otros"

@app.route("/api/inferir-categoria", methods=["GET"])
def inferir_categoria_api():
    invima_id = request.args.get("invima_id", "")
    principio = request.args.get("principio", "")
    nombre = request.args.get("nombre", "")

    if invima_id:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT categoria, producto FROM invima WHERE id = %s", (invima_id,))
        row = cur.fetchone()
        conn.close()
        if row:
            cat_bd = row["categoria"] or ""
            if cat_bd:
                return jsonify({"categoria": cat_bd, "fuente": "invima"})
            return jsonify({"categoria": inferir_categoria(row["producto"], row["producto"]), "fuente": "inferred"})

    return jsonify({"categoria": inferir_categoria(principio, nombre), "fuente": "keyword"})

@app.route("/api/categorias", methods=["GET"])
def categorias_listar():
    return jsonify(CATEGORIAS_PREDEFINIDAS)

# ── CATÁLOGO PROPIO ─────────────────────────────────────
@app.route("/api/catalogo", methods=["GET"])
def catalogo_listar():
    conn = get_db()
    rows = query(conn, "SELECT * FROM catalogo ORDER BY nombre").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/catalogo", methods=["POST"])
def catalogo_crear():
    d = request.json
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO catalogo (nombre, principio, laboratorio, descripcion)
        VALUES (%s, %s, %s, %s)
        RETURNING id
    """, (d["nombre"], d.get("principio",""), d.get("laboratorio",""), d.get("descripcion","")))
    nuevo_id = cur.fetchone()["id"]
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": nuevo_id})

# ── INVENTARIO ───────────────────────────────────────────
@app.route("/api/inventario", methods=["GET"])
def inventario_listar():
    conn = get_db()
    rows = query(conn, """
        SELECT * FROM inventario ORDER BY nombre
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/inventario", methods=["POST"])
def inventario_agregar():
    d = request.json
    conn = get_db()
    existing = query(conn, """
        SELECT id, cantidad FROM inventario
        WHERE (invima_id = %s AND invima_id IS NOT NULL)
           OR (catalogo_id = %s AND catalogo_id IS NOT NULL)
    """, (d.get("invima_id"), d.get("catalogo_id"))).fetchone()

    if existing:
        query(conn, """
            UPDATE inventario SET cantidad = cantidad + %s, precio = %s
            WHERE id = %s
        """, (d["cantidad"], d["precio"], existing["id"]))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "accion": "actualizado"})
    else:
        categoria = d.get("categoria", "") or inferir_categoria(d.get("principio", ""), d.get("nombre", ""))
        query(conn, """
            INSERT INTO inventario
                (invima_id, catalogo_id, nombre, principio, laboratorio, registro, cantidad, precio, fecha_vencimiento, lote, categoria)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            d.get("invima_id"), d.get("catalogo_id"),
            d["nombre"], d.get("principio",""), d.get("laboratorio",""),
            d.get("registro",""), d["cantidad"], d["precio"],
            d.get("fecha_vencimiento",""), d.get("lote",""), categoria
        ))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "accion": "creado"})

@app.route("/api/inventario/<int:item_id>", methods=["DELETE"])
def inventario_eliminar(item_id):
    conn = get_db()
    query(conn, "DELETE FROM inventario WHERE id = %s", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/inventario/<int:item_id>", methods=["PUT"])
def inventario_actualizar(item_id):
    d = request.json
    conn = get_db()
    query(conn, """
        UPDATE inventario SET 
            precio = %s, cantidad = %s, fecha_vencimiento = %s, lote = %s, categoria = %s
        WHERE id = %s
    """, (d.get("precio"), d.get("cantidad"), d.get("fecha_vencimiento"), d.get("lote"), d.get("categoria", ""), item_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/inventario/importar", methods=["POST"])
def inventario_importar():
    import pandas as pd
    from werkzeug.utils import secure_filename
    
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400
    
    try:
        df = pd.read_excel(file)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al leer Excel: {str(e)}"}), 400
    
    COLUMNAS_VALIDAS = ["nombre", "principio", "laboratorio", "cantidad", "precio", "fecha_vencimiento", "lote", "categoria"]
    COLUMNAS_ALIAS = {
        "producto": "nombre", "medicamento": "nombre", "articulo": "nombre",
        "componente": "principio", "genérico": "principio",
        "marca": "laboratorio", "fabricante": "laboratorio",
        "stock": "cantidad", "existencias": "cantidad",
        "valor": "precio", "costo": "precio",
        "vencimiento": "fecha_vencimiento", "caduca": "fecha_vencimiento",
        "lote": "lote", "categoría": "categoria",
        "precio unitario (cop)": "precio",
        "precio unitario": "precio",
        "número de unidades": "cantidad",
        "numero de unidades": "cantidad",
        "unidades": "cantidad"
    }
    
    df.columns = [c.strip().lower() for c in df.columns]
    df.rename(columns=COLUMNAS_ALIAS, inplace=True)
    
    cols_disponibles = [c for c in df.columns if c in COLUMNAS_VALIDAS]
    df = df[cols_disponibles]
    
    for col in ["cantidad", "precio"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    conn = get_db()
    preview = []
    
    for idx, row in df.iterrows():
        nombre_raw = row.get("nombre")
        if pd.isna(nombre_raw) or (isinstance(nombre_raw, float) and str(nombre_raw) == "nan"):
            continue
        nombre = str(nombre_raw).strip()
        if not nombre:
            continue
        
        invima_match = None
        if nombre:
            invima_match = query(conn, """
                SELECT id, principio_activo, titular, registro
                FROM invima
                WHERE producto ILIKE %s OR principio_activo ILIKE %s
                LIMIT 1
            """, (f"%{nombre}%", f"%{nombre}%")).fetchone()
        
        item = {
            "nombre": nombre,
            "principio": row.get("principio", "") or (invima_match["principio_activo"] if invima_match else ""),
            "laboratorio": row.get("laboratorio", "") or (invima_match["titular"] if invima_match else ""),
            "registro": invima_match["registro"] if invima_match else "",
            "cantidad": int(row.get("cantidad", 0)),
            "precio": float(row.get("precio", 0)),
            "fecha_vencimiento": str(row.get("fecha_vencimiento", ""))[:10] if pd.notna(row.get("fecha_vencimiento")) else "",
            "lote": str(row.get("lote", "")) if pd.notna(row.get("lote")) else "",
            "categoria": str(row.get("categoria", "")) if pd.notna(row.get("categoria")) else inferir_categoria(row.get("principio", ""), nombre),
            "invima_id": invima_match["id"] if invima_match else None,
            "match": invima_match is not None
        }
        preview.append(item)
    
    conn.close()
    return jsonify({"ok": True, "preview": preview, "columnas": cols_disponibles})

@app.route("/api/inventario/confirmar-importacion", methods=["POST"])
def inventario_confirmar_importacion():
    d = request.json
    items = d.get("items", [])
    
    if not items:
        return jsonify({"ok": False, "error": "No hay items"}), 400
    
    conn = get_db()
    insertados = 0
    actualizados = 0
    
    for item in items:
        nombre = item.get("nombre", "").strip()
        if not nombre:
            continue
        
        existing = query(conn, """
            SELECT id, cantidad FROM inventario
            WHERE (invima_id = %s AND invima_id IS NOT NULL)
               OR (nombre = %s)
        """, (item.get("invima_id"), nombre)).fetchone()
        
        if existing:
            query(conn, """
                UPDATE inventario SET cantidad = cantidad + %s, precio = %s
                WHERE id = %s
            """, (item.get("cantidad", 0), item.get("precio", 0), existing["id"]))
            actualizados += 1
        else:
            query(conn, """
                INSERT INTO inventario (invima_id, nombre, principio, laboratorio, registro, cantidad, precio, fecha_vencimiento, lote, categoria)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                item.get("invima_id"), nombre, item.get("principio", ""), item.get("laboratorio", ""),
                item.get("registro", ""), item.get("cantidad", 0), item.get("precio", 0),
                item.get("fecha_vencimiento", ""), item.get("lote", ""), item.get("categoria", "")
            ))
            insertados += 1
    
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "insertados": insertados, "actualizados": actualizados})

@app.route("/api/inventario/recategorizar", methods=["POST"])
def inventario_recategorizar():
    conn = get_db()
    productos = query(conn, "SELECT id, principio, nombre FROM inventario WHERE categoria IS NULL OR categoria = ''").fetchall()
    
    actualizados = 0
    for p in productos:
        nueva_cat = inferir_categoria(p["principio"], p["nombre"])
        if nueva_cat:
            query(conn, "UPDATE inventario SET categoria = %s WHERE id = %s", (nueva_cat, p["id"]))
            actualizados += 1
    
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "actualizados": actualizados})

# ── REPORTES ───────────────────────────────────────────────
@app.route("/api/reportes/inventario")
def reporte_inventario():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime
    
    conn = get_db()
    rows = query(conn, """
        SELECT nombre, laboratorio, cantidad, precio, fecha_vencimiento, categoria
        FROM inventario WHERE cantidad > 0 ORDER BY cantidad ASC
    """).fetchall()
    conn.close()
    
    data = [dict(r) for r in rows]
    df = pd.DataFrame(data)
    
    if not df.empty:
        df["valor_total"] = df["cantidad"] * df["precio"]
        df["alerta_stock"] = df["cantidad"].apply(lambda x: "CRÍTICO" if x <= 5 else ("BAJO" if x <= 10 else "OK"))
        df = df[["nombre", "laboratorio", "cantidad", "precio", "fecha_vencimiento", "categoria", "valor_total", "alerta_stock"]]
        df.columns = ["Nombre", "Laboratorio", "Cantidad", "Precio", "Fecha Vencimiento", "Categoría", "Valor Total", "Alerta Stock"]
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Inventario", index=False)
        
        ws = writer.sheets["Inventario"]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, col in enumerate(ws[1], start=1):
            col.fill = header_fill
            col.font = header_font
            col.alignment = header_alignment
        
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        critico_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
        critico_font = Font(color="990000")
        bajo_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        bajo_font = Font(color="7D6608")
        ok_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        ok_font = Font(color="274E13")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = grey_fill
            
            alerta_cell = row[7]
            alerta_val = alerta_cell.value
            if alerta_val == "CRÍTICO":
                alerta_cell.fill = critico_fill
                alerta_cell.font = critico_font
            elif alerta_val == "BAJO":
                alerta_cell.fill = bajo_fill
                alerta_cell.font = bajo_font
            elif alerta_val == "OK":
                alerta_cell.fill = ok_fill
                alerta_cell.font = ok_font
        
        last_data_row = ws.max_row + 1
        ws.cell(row=last_data_row, column=1, value="Total productos").font = Font(bold=True)
        ws.cell(row=last_data_row, column=3, value=len(df)).font = Font(bold=True)
        ws.cell(row=last_data_row, column=3).number_format = "#,##0"
        
        ws.cell(row=last_data_row + 1, column=1, value="Total unidades").font = Font(bold=True)
        ws.cell(row=last_data_row + 1, column=3, value=int(df["Cantidad"].sum())).font = Font(bold=True)
        ws.cell(row=last_data_row + 1, column=3).number_format = "#,##0"
        
        price_col = ws["D"]
        for cell in price_col[1:]:
            cell.number_format = "$#,##0"
        
        total_col = ws["G"]
        for cell in total_col[1:]:
            cell.number_format = "$#,##0"
        
        qty_col = ws["C"]
        for cell in qty_col[1:]:
            cell.number_format = "#,##0"
        
        date_col = ws["E"]
        for cell in date_col[1:]:
            cell.number_format = "DD/MM/YYYY"
        
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 4
        
        if not df.empty:
            resumen = pd.DataFrame({
                "Métrica": ["Total productos", "Total unidades", "Valor total stock", "Bajo stock", "Crítico"],
                "Valor": [
                    len(df),
                    df["Cantidad"].sum() if "Cantidad" in df else 0,
                    df["Valor Total"].sum() if "Valor Total" in df else 0,
                    len(df[df["Cantidad"] <= 10]) if "Cantidad" in df else 0,
                    len(df[df["Cantidad"] <= 5]) if "Cantidad" in df else 0
                ]
            })
            resumen.to_excel(writer, sheet_name="Resumen", index=False)
    
    output.seek(0)
    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"inventario_{datetime.now().strftime(FMT_FECHA_REPORTE)}.xlsx"
    )

@app.route("/api/reportes/ventas")
def reporte_ventas():
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime
    
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    
    conn = get_db()
    sql = "SELECT nombre, laboratorio, cantidad, precio_unitario, total, fecha FROM ventas WHERE 1=1"
    params = []
    if desde:
        sql += " AND fecha::date >= %s::date"
        params.append(desde)
    if hasta:
        sql += " AND fecha::date <= %s::date"
        params.append(hasta)
    sql += " ORDER BY fecha DESC"
    
    rows = query(conn, sql, params).fetchall()
    conn.close()
    
    data = [dict(r) for r in rows]
    df = pd.DataFrame(data)
    
    if not df.empty:
        df.columns = ["Nombre", "Laboratorio", "Cantidad", "Precio Unitario", "Total", "Fecha"]
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Ventas", index=False)
        
        ws = writer.sheets["Ventas"]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in ws[1]:
            col.fill = header_fill
            col.font = header_font
            col.alignment = header_alignment
        
        grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = grey_fill
        
        price_col = ws["D"]
        for cell in price_col[1:]:
            cell.number_format = "$#,##0"
        
        total_col = ws["E"]
        for cell in total_col[1:]:
            cell.number_format = "$#,##0"
        
        qty_col = ws["C"]
        for cell in qty_col[1:]:
            cell.number_format = "#,##0"
        
        date_col = ws["F"]
        for cell in date_col[1:]:
            cell.number_format = "DD/MM/YYYY"
        
        name_col = ws["A"]
        lab_col = ws["B"]
        for cell in name_col[1:]:
            cell.alignment = Alignment(horizontal="left")
        for cell in lab_col[1:]:
            cell.alignment = Alignment(horizontal="left")
        
        for column in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = max_length + 4
        
        if not df.empty:
            last_row = ws.max_row + 1
            ws.cell(row=last_row, column=1, value="Total ventas").font = Font(bold=True)
            ws.cell(row=last_row, column=5, value=int(df["Total"].sum())).font = Font(bold=True)
            ws.cell(row=last_row, column=5).number_format = "$#,##0"
    
    output.seek(0)
    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"ventas_{desde or 'all'}_{hasta or 'all'}.xlsx"
    )

@app.route("/api/reportes/utilidad")
def reporte_utilidad():
    import pandas as pd
    from io import BytesIO
    from datetime import datetime
    
    desde = request.args.get("desde", "")
    hasta = request.args.get("hasta", "")
    
    conn = get_db()
    sql = "SELECT * FROM ventas WHERE 1=1"
    params = []
    if desde:
        sql += " AND fecha::date >= %s::date"
        params.append(desde)
    if hasta:
        sql += " AND fecha::date <= %s::date"
        params.append(hasta)
    sql += " ORDER BY fecha DESC"
    
    ventas = query(conn, sql, params).fetchall()
    
    inventario_actual = query(conn, "SELECT SUM(cantidad * precio) as costo FROM inventario").fetchone()
    conn.close()
    
    ventas_data = [dict(r) for r in ventas]
    df_ventas = pd.DataFrame(ventas_data)
    
    ingresos = df_ventas["total"].sum() if not df_ventas.empty else 0
    costo_inventario = inventario_actual["costo"] or 0
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not df_ventas.empty:
            df_ventas.to_excel(writer, sheet_name="Ventas", index=False)
        
        resumen = pd.DataFrame({
            "Concepto": ["Ingresos por ventas", "Costo estimado inventario", "Margen bruto estimado"],
            "Valor": [ingresos, costo_inventario, ingresos - costo_inventario]
        })
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
    
    output.seek(0)
    from flask import send_file
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"utilidad_{datetime.now().strftime(FMT_FECHA_REPORTE)}.xlsx"
    )

@app.route("/api/dashboard/alertas", methods=["GET"])
def dashboard_alertas():
    import datetime
    conn = get_db()
    
    fecha_limite = (datetime.datetime.now() + datetime.timedelta(days=DIAS_ALERTA_VENCIMIENTO)).strftime(DATE_FMT)
    
    bajo_stock = query(conn, """
        SELECT nombre, cantidad,
               CASE WHEN cantidad <= %s THEN 'critico' ELSE 'bajo' END as nivel
        FROM inventario
        WHERE cantidad <= %s
        ORDER BY cantidad ASC
    """, (STOCK_CRITICO, STOCK_MINIMO)).fetchall()
    
    vencer = query(conn, """
        SELECT nombre, fecha_vencimiento, cantidad
        FROM inventario
        WHERE fecha_vencimiento IS NOT NULL AND fecha_vencimiento != '' AND fecha_vencimiento <= %s
        ORDER BY fecha_vencimiento ASC
    """, (fecha_limite,)).fetchall()
    
    conn.close()
    return jsonify({
        "bajo_stock": [dict(r) for r in bajo_stock],
        "proximos_vencer": [dict(r) for r in vencer]
    })

@app.route("/api/dashboard/ventas-hoy", methods=["GET"])
def dashboard_ventas_hoy():
    conn = get_db()
    row = query(conn, """
        SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as transacciones
        FROM ventas
        WHERE fecha::date = CURRENT_DATE
    """).fetchone()
    conn.close()
    return jsonify({"total": row["total"], "transacciones": row["transacciones"]})

@app.route("/api/dashboard", methods=["GET"])
def dashboard():
    import datetime
    conn = get_db()

    total = query(conn, "SELECT SUM(cantidad) as total FROM inventario").fetchone()["total"] or 0
    total_productos = query(conn, "SELECT COUNT(*) as count FROM inventario WHERE cantidad > 0").fetchone()["count"]
    bajo_stock = query(conn, "SELECT COUNT(*) as count FROM inventario WHERE cantidad <= %s", (STOCK_MINIMO,)).fetchone()["count"]

    fecha_limite = (datetime.datetime.now() + datetime.timedelta(days=DIAS_ALERTA_VENCIMIENTO)).strftime(DATE_FMT)
    vencer = query(conn, """
        SELECT COUNT(*) as count FROM inventario
        WHERE fecha_vencimiento IS NOT NULL AND fecha_vencimiento != '' AND fecha_vencimiento <= %s
    """, (fecha_limite,)).fetchone()["count"]

    ventas_mes = query(conn, """
        SELECT COALESCE(SUM(total), 0) as total
        FROM ventas
        WHERE fecha::date >= DATE_TRUNC('month', CURRENT_DATE)::date
    """).fetchone()["total"] or 0

    conn.close()
    return jsonify({
        "total_productos": total_productos,
        "total_unidades": total,
        "bajo_stock": bajo_stock,
        "prox_vencer": vencer,
        "ventas_mes": ventas_mes
    })

@app.route("/api/dashboard/ventas-semana", methods=["GET"])
def dashboard_ventas_semana():
    import datetime
    conn = get_db()
    rows = query(conn, """
        SELECT fecha::date as fecha, COALESCE(SUM(total), 0) as total
        FROM ventas
        WHERE fecha::date >= CURRENT_DATE - INTERVAL '6 days'
        GROUP BY fecha::date
        ORDER BY fecha::date ASC
    """).fetchall()
    conn.close()
    
    existentes = {str(r["fecha"]): r["total"] for r in rows}
    hoy = datetime.date.today()
    return jsonify([
        {
            "fecha": (hoy - datetime.timedelta(days=i)).isoformat(),
            "total": existentes.get((hoy - datetime.timedelta(days=i)).isoformat(), 0)
        }
        for i in range(6, -1, -1)
    ])

@app.route("/api/dashboard/top-productos", methods=["GET"])
def dashboard_top_productos():
    conn = get_db()
    rows = query(conn, """
        SELECT nombre, SUM(cantidad) as vendidos
        FROM ventas
        GROUP BY nombre
        ORDER BY vendidos DESC
        LIMIT 5
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

# ── VENTAS ───────────────────────────────────────────────
@app.route("/api/ventas", methods=["GET"])
def ventas_listar():
    conn = get_db()
    rows = query(conn, """
        SELECT * FROM ventas ORDER BY fecha DESC LIMIT 100
    """).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/ventas", methods=["POST"])
def registrar_venta():
    d = request.json
    conn = get_db()
    inv_id = d.get("inventario_id")
    if not inv_id:
        conn.close()
        return jsonify({"ok": False, "error": "inventario_id requerido"}), 400

    item = query(conn, """
        SELECT id, nombre, laboratorio, cantidad, precio FROM inventario WHERE id = %s
    """, (inv_id,)).fetchone()

    if not item:
        conn.close()
        return jsonify({"ok": False, "error": "Producto no encontrado"}), 404

    if item["cantidad"] < d["cantidad"]:
        conn.close()
        return jsonify({"ok": False, "error": "Stock insuficiente"}), 400

    total = item["precio"] * d["cantidad"]

    query(conn, """
        UPDATE inventario SET cantidad = cantidad - %s WHERE id = %s
    """, (d["cantidad"], item["id"]))

    import datetime
    fecha_actual = datetime.datetime.now().isoformat(timespec="seconds")
    query(conn, """
        INSERT INTO ventas (inventario_id, nombre, laboratorio, cantidad, precio_unitario, total, fecha)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (item["id"], item["nombre"], item["laboratorio"], d["cantidad"], item["precio"], total, fecha_actual))

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "total": total})

# ── ANALÍTICAS (PREMIUM) ───────────────────────────────────
def _premium_required():
    if not PREMIUM_ENABLED:
        return jsonify({"ok": False, "premium_required": True, "error": "Función premium"}), 403
    return None


def _variacion_pct(actual, anterior):
    if not anterior:
        return None
    return round(((actual - anterior) / anterior) * 100, 1)


@app.route("/api/analytics/comparativa", methods=["GET"])
def analytics_comparativa():
    if err := _premium_required():
        return err
    conn = get_db()
    actual = query(conn, """
        SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as transacciones
        FROM ventas
        WHERE fecha::date >= DATE_TRUNC('month', CURRENT_DATE)::date
    """).fetchone()
    anterior = query(conn, """
        SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as transacciones
        FROM ventas
        WHERE fecha::date >= (DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '1 month')::date
          AND fecha::date <  DATE_TRUNC('month', CURRENT_DATE)::date
    """).fetchone()
    conn.close()
    a_total, a_tx = actual["total"] or 0, actual["transacciones"] or 0
    b_total, b_tx = anterior["total"] or 0, anterior["transacciones"] or 0
    return jsonify({
        "actual": {"total": a_total, "transacciones": a_tx, "ticket_promedio": (a_total / a_tx) if a_tx else 0},
        "anterior": {"total": b_total, "transacciones": b_tx, "ticket_promedio": (b_total / b_tx) if b_tx else 0},
        "variacion_total_pct": _variacion_pct(a_total, b_total),
        "variacion_transacciones_pct": _variacion_pct(a_tx, b_tx),
    })


@app.route("/api/analytics/rotacion", methods=["GET"])
def analytics_rotacion():
    if err := _premium_required():
        return err
    conn = get_db()
    rows = query(conn, """
        SELECT i.nombre,
               i.cantidad as stock,
               COALESCE(SUM(v.cantidad), 0) as vendidos,
               CASE WHEN i.cantidad > 0
                    THEN CAST(COALESCE(SUM(v.cantidad), 0) AS REAL) / i.cantidad
                    ELSE 0 END as rotacion
        FROM inventario i
        LEFT JOIN ventas v ON v.inventario_id = i.id
        WHERE i.cantidad > 0
        GROUP BY i.id
        ORDER BY rotacion DESC
        LIMIT 10
    """).fetchall()
    conn.close()
    return jsonify({"productos": [dict(r) for r in rows]})


@app.route("/api/analytics/rentabilidad", methods=["GET"])
def analytics_rentabilidad():
    if err := _premium_required():
        return err
    conn = get_db()
    rows = query(conn, """
        SELECT i.categoria, COALESCE(SUM(v.total), 0) as ventas
        FROM ventas v
        JOIN inventario i ON v.inventario_id = i.id
        WHERE i.categoria IS NOT NULL AND i.categoria != ''
        GROUP BY i.categoria
        ORDER BY ventas DESC
    """).fetchall()
    conn.close()
    resultado = []
    for r in rows:
        margen = MARGEN_POR_CATEGORIA.get(r["categoria"], 0.30)
        ventas = r["ventas"] or 0
        resultado.append({
            "categoria": r["categoria"],
            "ventas": ventas,
            "utilidad": ventas * margen,
            "margen_pct": round(margen * 100, 1),
        })
    return jsonify({"categorias": resultado})


@app.route("/api/analytics/margen", methods=["GET"])
def analytics_margen():
    if err := _premium_required():
        return err
    conn = get_db()
    rows = query(conn, """
        SELECT i.categoria, COALESCE(SUM(v.total), 0) as ventas
        FROM ventas v
        JOIN inventario i ON v.inventario_id = i.id
        WHERE i.categoria IS NOT NULL AND i.categoria != ''
        GROUP BY i.categoria
    """).fetchall()
    conn.close()
    total_ventas = 0
    total_utilidad = 0
    for r in rows:
        margen = MARGEN_POR_CATEGORIA.get(r["categoria"], 0.30)
        v = r["ventas"] or 0
        total_ventas += v
        total_utilidad += v * margen
    margen_pct = round((total_utilidad / total_ventas) * 100, 1) if total_ventas else 0
    return jsonify({
        "total_ventas": total_ventas,
        "total_utilidad": total_utilidad,
        "margen_pct": margen_pct,
    })


@app.route("/api/analytics/proyeccion", methods=["GET"])
def analytics_proyeccion():
    import datetime
    if err := _premium_required():
        return err
    hoy = datetime.date.today()
    if hoy.month == 12:
        fin_mes = hoy.replace(day=31)
    else:
        fin_mes = (hoy.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)
    dias_transcurridos = hoy.day
    dias_totales = fin_mes.day
    dias_restantes = (fin_mes - hoy).days

    conn = get_db()
    ventas_mes = query(conn, """
        SELECT COALESCE(SUM(total), 0) as total
        FROM ventas
        WHERE fecha::date >= DATE_TRUNC('month', CURRENT_DATE)::date
    """).fetchone()["total"] or 0
    conn.close()

    promedio_diario = ventas_mes / dias_transcurridos if dias_transcurridos > 0 else 0
    proyeccion = promedio_diario * dias_totales
    return jsonify({
        "ventas_mes_actual": ventas_mes,
        "dias_transcurridos": dias_transcurridos,
        "dias_restantes": dias_restantes,
        "dias_totales": dias_totales,
        "promedio_diario": promedio_diario,
        "proyeccion": proyeccion,
    })


# ── PÁGINA PRINCIPAL ─────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html", premium_enabled=PREMIUM_ENABLED)

# ── INIT_DB EN TOP-LEVEL ────────────────────────────────
# ¿Por qué aquí y no dentro de `if __name__ == "__main__"`?
# ------------------------------------------------------------
# Render ejecuta `gunicorn app:app` (definido en Procfile).
# gunicorn IMPORTA el módulo `app` y busca la variable `app` (la Flask instance).
# NO ejecuta el bloque `if __name__ == "__main__"`, porque ese solo corre
# cuando haces `python app.py` directamente.
#
# Si init_db() quedara dentro de ese bloque, en producción las tablas NUNCA
# se crearían y todos los endpoints que tocan la BD darían 500.
#
# ¿Por qué es seguro ejecutarlo al importar?
# - init_db() usa `CREATE TABLE IF NOT EXISTS` y `INSERT ... ON CONFLICT`
#   (en este código, el sembrado de 158 productos INVIMA usa un check
#   `SELECT COUNT(*) FROM invima` antes de insertar). Es idempotente.
# - Aunque gunicorn tenga N workers, init_db() corre 1 vez por worker.
#   El costo es despreciable (158 inserts en una BD ya creada se saltan).
#
# Si en el futuro quieres volver al modo "solo en python app.py":
#   1. Mueve init_db() de vuelta dentro del bloque `if __name__ == "__main__"`
#   2. Crea un script separado (ej: scripts/init_db.py) que ejecute init_db()
#   3. Agrega un paso en el "Build Command" de Render:
#      pip install -r requirements.txt && python scripts/init_db.py
#   4. El Procfile queda con `gunicorn app:app` (sin init_db)
# Eso es más limpio para apps grandes con muchas tablas/migraciones.
# Para esta app (4 tablas, 158 productos), el approach top-level es suficiente.
init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
